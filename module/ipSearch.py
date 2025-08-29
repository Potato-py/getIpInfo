import os,sys,re
from configparser import ConfigParser
import requests,time,json
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
from openpyxl import load_workbook

try:
    from font import *
except:
    from module.font import *
Processing=str(Processing())
Information=str(Information())
Detected=str(Detected())
Result=str(Result())
Error=str(Error())

try:
    from qqwry import QQwry
    from qqwry import updateQQwry
except:
    print("pip install qqwry-py3")

#更新纯真ip数据库,可手动触发
def updatePos():
    print(Processing+"正在更新最新纯真IP数据库，请稍等……")
    updateQQwry('qqwry.dat')
    print(Result+"纯真IP数据库更新完毕！")

def ipPos(ipList):  #返回字典、自带去重
    q = QQwry()
    # 判断是否存在纯真ip数据库
    if not os.path.exists('qqwry.dat'):
        updatePos()
    # 加载纯真ip数据库
    q.load_file('qqwry.dat')
    # 查询ip归属地
    result = {}
    for ip in ipList:
        pos = q.lookup(ip)[0]
        result[ip] = pos
    return result

def initCookie(needNewCookie=False):
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.ini")
    cf = ConfigParser()
    try:
        cf.read(file_path, encoding='utf-8')
        session = cf.get('arsenal', 'cookie-Session')
        if(needNewCookie):
            a=cf+1#故意抛异常
    except:
        session = input('\n'+Information+bold("请登录https://ti.qianxin.com/后，在此输入Cookie中session的值：\n")+Input())
        cf.add_section('arsenal') if not cf.has_section('arsenal') else 0
        cf.set('arsenal', 'cookie-Session', session)
        cf.write(open(file_path, 'w+'))
    return session

#默认会输出Result目录下，analysis控制是否解析xlsx并返回结果（默认不解析）
#tempDataList别太大，200以内
def ipReputationFromQax(tempDataList,analysis=False,needNewCookie=False):
    dataList=[]
    for data in tempDataList:
        if not re.match("^(127\\.0\\.0\\.1)|(0\\.0\\.0\\.0)|(localhost)|(10\\.\\d{1,3}\\.\\d{1,3}\\.\\d{1,3})|(172\\.((1[6-9])|(2\\d)|(3[01]))\\.\\d{1,3}\\.\\d{1,3})|(192\\.168\\.\\d{1,3}\\.\\d{1,3})$",data):
            dataList.append(data)
    if(len(dataList)==0):
        print(Error+"传入无外网地址")
        print(tempDataList)
    session = initCookie(needNewCookie)
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.111 Safari/537.36",
        "Connection": "close",
        "Cookie": f"session={session}",
        "Content-Type": "application/json;charset=UTF-8",
        "Referer": "https://ares.ti.qianxin.com/ares/tools/ip-reputation"
    }
    data = '{"upload_type_name":"ip_reputation_analysis","file_str":"'+r"\n".join(dataList)+'"}'
    msg = ''
    try:
        req1 = requests.post(url="https://ares.ti.qianxin.com/arsenal/api/v2/uploadStr", headers=headers, data=data, verify=False, allow_redirects=False,timeout=5)
        msg = req1.json()['message']
    except Exception as e:
        print(Error+str(e))
    if msg == 'no access to this module':
        print(Error+"Cookie失效，请更新您的Cookie-session的值")
        ipReputationFromQax(dataList,analysis,True)
        return False
    if msg == '\u4eca\u65e5\u514d\u8d39\u8bd5\u7528\u6b21\u6570\u5df2\u7ecf\u4f7f\u7528\u5b8c\u6bd5':
        print(Error+msg)
        print(Information+"基于接口bug，请更换您的ip地址即可绕过")
        return False
    if(msg=='' or msg != 'upload task success'):
        print(Error+msg)
        print(Information+"请检查接口是否存在异常/更换您的cookie值")
        return False
    time.sleep(3)#等待服务端解析
    data = '{"page": 1,"page_size": 10}'
    try:
        req1 = requests.post(url="https://ares.ti.qianxin.com/arsenal/api/v2/task-list", headers=headers, data=data, verify=False, allow_redirects=False,timeout=5)
        msg = req1.json()['data']['result'][0]['result_file_url']
        fileName = req1.json()['data']['result'][0]['result_file_name']
        if(msg == None or msg==""):
            print(Error+req1.json()['message'])
            return False
    except Exception as e:
        print(Error+str(e))
    try:
        req1 = requests.get(url=msg, verify=False, allow_redirects=False,timeout=5)
        if not os.path.exists('./Result'):
            os.makedirs('./Result')
        with open(f'./Result/{fileName}', 'wb') as f:
            f.write(req1.content)
    except Exception as e:
        print(Error+str(e))

    if(not analysis):
        print(Result+f'详细内容已输出至文件：./Result/{fileName}')
        return False
    #解析xlsx文档，读取数据
    result=[]
    wb = load_workbook(f'./Result/{fileName}', data_only=True)
    sheetNames = wb.sheetnames
    for i in sheetNames:
        sheet = wb[i]
        sheetData = []
        for row in sheet.iter_rows(values_only=True):
            sheetData.append(list(row))
        result.append(sheetData)

    printDataList = result[0][1:]
    printT( [15,10,7,7,14,22,5] ,"top")
    printT( [["IP",15],["国家",10],["归属组织",7],["解析域名",14],["恶意类型",22],["...",5]])
    for printData in printDataList:
        printT( [15,10,7,7,14,22,5] ,"middle")
        type ='-' if printData[-4]=="" else printData[-4]
        domain = '-' if printData[13]=="" else printData[13]
        printT( [[printData[0],15],[printData[2],10],[printData[10],7],[domain,14],[type,22],["...",5]])
    printT( [15,10,7,7,14,22,5] ,"bottom")
    print(Result+f'IP详细内容已输出至文件：./Result/{fileName}')
    return result

# dataList=['103.85.84.160','46.19.138.162','113.56.96.34']
# posDict = ipPos(dataList)
# print(posDict)
# ipInfo = ipReputationFromQax(dataList,True)
# print(ipInfo)